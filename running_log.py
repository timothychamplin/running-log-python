from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

FILE_NAME = "running_log.xlsx"


def get_run_input():
    """Ask the user for information about their run."""

    date = input("Date of run (YYYY-MM-DD, or press Enter for today): ")

    if date.strip() == "":
        date = datetime.today().strftime("%Y-%m-%d")

    while True:
        try:
            distance = float(input("Distance in miles: "))
            if distance <= 0:
                print("Distance must be greater than 0.")
            else:
                break
        except ValueError:
            print("Please enter a valid number for distance.")

    total_time = get_valid_time()

    pace = calculate_pace(distance, total_time)

    run_type = input("Run type (easy, workout, long run, race, recovery): ")
    location = input("Location: ")
    shoes = input("Shoes worn: ")

    while True:
        try:
            feel = int(input("How did you feel? 1-10: "))
            if 1 <= feel <= 10:
                break
            else:
                print("Please enter a number from 1 to 10.")
        except ValueError:
            print("Please enter a whole number.")

    notes = input("Notes: ")

    return [
        date,
        distance,
        total_time,
        pace,
        run_type,
        location,
        shoes,
        feel,
        notes
    ]


def get_valid_time():
    """
    Ask for total run time.
    Format should be HH:MM:SS or MM:SS.
    """

    while True:
        total_time = input("Total time (HH:MM:SS or MM:SS): ")

        parts = total_time.split(":")

        if len(parts) == 2:
            minutes, seconds = parts
            if minutes.isdigit() and seconds.isdigit():
                return total_time

        elif len(parts) == 3:
            hours, minutes, seconds = parts
            if hours.isdigit() and minutes.isdigit() and seconds.isdigit():
                return total_time

        print("Invalid time format. Example: 25:30 or 1:05:30")


def time_to_seconds(time_string):
    """Convert a time string into total seconds."""

    parts = time_string.split(":")

    if len(parts) == 2:
        minutes = int(parts[0])
        seconds = int(parts[1])
        return minutes * 60 + seconds

    elif len(parts) == 3:
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = int(parts[2])
        return hours * 3600 + minutes * 60 + seconds


def calculate_pace(distance, total_time):
    """Calculate pace per mile based on distance and total time."""

    total_seconds = time_to_seconds(total_time)
    pace_seconds = total_seconds / distance

    minutes = int(pace_seconds // 60)
    seconds = int(pace_seconds % 60)

    return f"{minutes}:{seconds:02d} per mile"


def save_to_excel(run_data):
    """Save the run data to an Excel file."""

    headers = [
        "Date",
        "Distance",
        "Total Time",
        "Pace",
        "Run Type",
        "Location",
        "Shoes",
        "Feel 1-10",
        "Notes"
    ]

    try:
        workbook = load_workbook(FILE_NAME)
        sheet = workbook.active

    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Running Log"
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

    sheet.append(run_data)

    adjust_column_widths(sheet)

    workbook.save(FILE_NAME)


def adjust_column_widths(sheet):
    """Make Excel columns wider so the file looks cleaner."""

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 2


def show_weekly_mileage():
    """Show mileage from the last 7 days."""

    try:
        workbook = load_workbook(FILE_NAME)
        sheet = workbook.active
    except FileNotFoundError:
        return

    today = datetime.today().date()
    week_ago = today - timedelta(days=7)

    total_miles = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        date_value = row[0]
        distance = row[1]

        try:
            run_date = datetime.strptime(str(date_value), "%Y-%m-%d").date()

            if week_ago <= run_date <= today:
                total_miles += float(distance)

        except:
            continue

    print(f"Your mileage for the last 7 days is: {total_miles:.2f} miles")


def main():
    """Main program loop."""

    print("Welcome to your Running Log!")

    while True:
        run_data = get_run_input()
        save_to_excel(run_data)

        print("\nRun saved successfully!")
        print(f"Saved to {FILE_NAME}")

        show_weekly_mileage()

        another = input("\nWould you like to enter another run? yes/no: ").lower()

        if another != "yes":
            print("Goodbye. Nice work getting the miles in!")
            break


main()